from zipfile import ZipFile
import os
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv

# – Запаковать в zip архив несколько разных файлов: pdf, xlsx, csv;
# – Положить его в ресурсы;
# – Реализовать чтение и проверку содержимого каждого файла из архива в виде тестов

# объявление переменных
path_ = 'resources'  # директория, в которую ляжет архив
# name_fale = 'names.csv' - для архивирования одного файла. не используется.
root_ = 'files'  # директория, в которой лежат файлы, которые нужно положить архив
n = os.listdir(root_)  # сохранение списка имен  фалов в директории
len_list: int = len(os.listdir(root_))  # сохранение длины списка файлов в директории
pdf_len = 0
pdf_text = "text"
value_xlsx = "text"
headers_csv = "text"
xlsx_size = 0
pdf_size = 0
cvs_size = 0
name_archiv = 'zip_1.zip'


def put_in_archive_folder(path_arch=path_, root_arch=root_):  # создаю архив, заношу файлы, сохраняю данные о файлах
    global zip_p, pdf_len, pdf_text, pdf_size, csv_size, headers_csv, xlsx_size, value_xlsx
    print('я начал создавать архив ' + name_archiv)
    zip_p = ZipFile(os.path.join(path_, name_archiv), 'w')  # открытие файла на запись
    # Запись одного файла в архив с указанием типа архивирования
    # zip_p.write((os.path.join(path_, name_fale)), compress_type=zipfile.ZIP_DEFLATED)
    # Запись нескольких файлов в архив
    for root, dirs, files in os.walk(root_):  # Список всех файлов и папок в директории folder
        for file in files:
            zip_p.write(os.path.join(root, file))  # Создание относительных путей и запись файлов в архив
            # print(file)
            if file.endswith(
                    '.pdf'):  # если файл в папке - запоминаю, сколько страниц и текст с первой страницы. и вес файла
                # print(file.title())
                reader = PdfReader(os.path.join(root, file.title()))
                pdf_size = os.path.getsize(os.path.join(root, file.title()))
                pdf_len = len(reader.pages)
                # print(len(reader.pages))
                pdf_page = reader.pages[0]
                pdf_text = pdf_page.extract_text()
                # print(text)
            if file.endswith('.csv'):  # запоминаю заголовки из csv и вес файла
                with open(os.path.join(root, file.title())) as csvfile:
                    csv_size = os.path.getsize(os.path.join(root, file.title()))
                    reader = csv.reader(csvfile)
                    headers_csv = str(next(reader))
            if file.endswith('.xlsx'):  # запоминаю данные в определенной ячейке. и вес файла
                xlsx_size = os.path.getsize(os.path.join(root, file.title()))
                workbook = load_workbook(os.path.join(root, file.title()))
                sheet = workbook.active
                value_xlsx = sheet.cell(row=3, column=2).value
            else:
                pass
    # print(zip_p.namelist())  # чтение содержимого
    list_zip: int = len(zip_p.namelist())  # сохранение длинны списка файлов в архиве
    zip_p.close()  # закрыть архив на запись
    print('я закончил создавать архив')
    return list_zip, zip_p, pdf_len, pdf_text, pdf_size, csv_size, headers_csv, xlsx_size, value_xlsx


# Проверки
def test_file_in_archive(len_list_zip, pdf_len, pdf_text, pdf_size, csv_size, headers_csv, xlsx_size, value_xlsx):
    print('я запустил тест')
    # папака содержит столько же файлов, сколько архив
    assert len_list == len_list_zip
    # имена файлов в архиве соответсвуют именам файлов в исходной папке
    for i in range(len(n)):  # итерация по i где i от 0 до длинны n
        s = str(zip_p.namelist()[i])  # беру имя с путем
        s = (s[s.find("/") + 1:])  # дорматирую имя с путем до names.csv
        # print(s)
        assert n[i] == s  # имя из списка файлов в папке == имени из списка файлов в архиве
    # Проверка соответсвия содержания
    for file in zip_p.namelist():
        if file.endswith('.pdf'):
            pdf_size_zip = os.path.getsize(file)
            reader_zip = PdfReader(file)
            assert pdf_len == len(reader_zip.pages)
            assert pdf_text in reader_zip.pages[0].extractText()
            assert pdf_size_zip == pdf_size
        if file.endswith('.csv'):
            csv_size_zip = os.path.getsize(file)
            reader = csv.reader(file)
            headers_csv_zip = str(next(reader))
            assert str(headers_csv[0]) == str(headers_csv_zip[0])
            assert csv_size_zip == csv_size
        if file.endswith('.xlsx'):
            xlsx_size_zip = os.path.getsize(file)
            workbook = load_workbook(file)
            sheet = workbook.active
            value_xlsx_zip = sheet.cell(row=3, column=2).value
            assert value_xlsx_zip == value_xlsx
            assert xlsx_size_zip == xlsx_size
        else:
            pass
        # удаление файла
    print('тест прошел успешно')
    # os.remove(os.path.join(path_, 'zip_1.zip'))
    # print('я удалил файл')


put_in_archive_folder()
test_file_in_archive(len_list_zip=len_list, pdf_len=pdf_len, pdf_text=pdf_text, pdf_size=pdf_size, csv_size=csv_size,
                     headers_csv=headers_csv, xlsx_size=xlsx_size, value_xlsx=value_xlsx)
